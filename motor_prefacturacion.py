import re
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import Font
 
ARCHIVO_PEDIDOS = "ISA INTERCOLOMBIA FEB.XLSX"
ARCHIVO_TARIFAS = "Valor kilo destino Colvanes 2026.xlsx"
SALIDA = ARCHIVO_PEDIDOS
 
HOJA_PEDIDOS = "MERCANCIA"
HOJA_TARIFAS = "DEFINITIVO 026"

COSTO_MANEJO_POR_UNIDAD = 6041
UMBRAL_DECLARADO_SEGURO = 750000
PORC_DESCUENTO = 0.25
PORC_SEGURO = 0.005
UMBRAL_SIMILAR = 5000



# Utilidades de texto

def normalizar(texto):
    if texto is None:
        return ""
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"\s+", " ", texto)
    return texto


def limpiar_ciudad(texto):
    t = normalizar(texto)
    if not t:
        return ""

    if t.strip() == "SANTA FE DE BOGOTA":
        t = "BOGOTA"

    ciudad, depto = t, ""

    m = re.match(r"^(.*?)\s*\((.*?)\)\s*$", t)
    if m:
        ciudad = (m.group(1) or "").strip()
        depto = (m.group(2) or "").strip()
    else:
        for sep in [" - ", "-", " / ", "/", " , ", ","]:
            if sep in t:
                partes = [p.strip() for p in t.split(sep, 1)]
                if len(partes) == 2:
                    ciudad, depto = partes[0], partes[1]
                break

    def limpiar_ruido(s: str) -> str:
        if not s:
            return ""
        s = s.replace("DISTRITO CAPITAL", " ")
        s = s.replace("DISTRITO ESPECIAL", " ")
        s = s.replace("D.C.", " ").replace("D C", " ").replace("DC", " ")
        s = s.replace("D.E.", " ").replace("D E", " ").replace("DE", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    ciudad = limpiar_ruido(ciudad)
    depto = limpiar_ruido(depto)

    if ciudad == "BOGOTA" or ciudad.startswith("BOGOTA "):
        return "BOGOTA"

    if depto:
        return f"{ciudad}-{depto}"

    return ciudad



# Utilidades numéricas

def _as_int(x):
    if isinstance(x, int):
        return x
    if isinstance(x, float):
        try:
            return int(x)
        except Exception:
            return 0
    if x is None:
        return 0
    try:
        s = str(x).strip()
        s = s.replace(".", "").replace(",", "").replace(" ", "")
        return int(float(s))
    except Exception:
        return 0


# Construcción de índices de la matriz de tarifas

def construir_indices(ws):

    origenes = {}
    destinos = {}

    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val:
            clave = limpiar_ciudad(val)
            origenes[clave] = col

    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 2).value
        if val:
            clave = limpiar_ciudad(val)
            destinos[clave] = row

    return origenes, destinos



# PROCESO MERCANCIA
def prefacturar():

    wb_pedidos = load_workbook(ARCHIVO_PEDIDOS)
    wb_tarifas = load_workbook(ARCHIVO_TARIFAS, data_only=True)

    ws_pedidos = wb_pedidos[HOJA_PEDIDOS]
    ws_tarifas = wb_tarifas[HOJA_TARIFAS]

    origenes, destinos = construir_indices(ws_tarifas)

    headers = {
        str(ws_pedidos.cell(1, c).value).strip().upper(): c
        for c in range(1, ws_pedidos.max_column + 1)
        if ws_pedidos.cell(1, c).value
    }

    col_origen   = headers["ORIGEN"]
    col_destino  = headers["DESTINO"]
    col_peso     = headers["PESO FACTURADO"]

    col_unidades  = headers.get("UNIDADES")
    col_declarado = headers.get("DECLARADO")
    col_total     = headers.get("TOTAL")

    base_col = ws_pedidos.max_column
    col_ciudad_origen   = base_col + 1
    col_ciudad_destino  = base_col + 2
    col_valor_kilo      = base_col + 3
    col_prefac          = base_col + 4
    col_obs             = base_col + 5
    col_comp_total      = base_col + 6

    titulos = {
        col_ciudad_origen:  "CIUDAD_ORIGEN",
        col_ciudad_destino: "CIUDAD_DESTINO",
        col_valor_kilo:     "VALOR_KILO",
        col_prefac:         "PREFAC_FLETE",
        col_obs:            "OBS_MATCH",
        col_comp_total:     "COMPARA_TOTAL",
    }

    for col, titulo in titulos.items():
        ws_pedidos.cell(1, col).value = titulo
        ws_pedidos.cell(1, col).font = Font(bold=True)

    for row in range(2, ws_pedidos.max_row + 1):

        origen_raw   = ws_pedidos.cell(row, col_origen).value
        destino_raw  = ws_pedidos.cell(row, col_destino).value
        peso         = _as_int(ws_pedidos.cell(row, col_peso).value)

        unidades = _as_int(ws_pedidos.cell(row, col_unidades).value if col_unidades else 0)
        declarado = _as_int(ws_pedidos.cell(row, col_declarado).value if col_declarado else 0)
        total_ext = _as_int(ws_pedidos.cell(row, col_total).value) if col_total else None

        origen  = limpiar_ciudad(origen_raw)
        destino = limpiar_ciudad(destino_raw)

        ws_pedidos.cell(row, col_ciudad_origen).value  = origen
        ws_pedidos.cell(row, col_ciudad_destino).value = destino

        col_o  = origenes.get(origen)
        fila_d = destinos.get(destino)

        if col_o and fila_d:
            raw_valor_kilo = ws_tarifas.cell(fila_d, col_o).value

            if raw_valor_kilo is None or (isinstance(raw_valor_kilo, str) and raw_valor_kilo.strip() == ""):
                ws_pedidos.cell(row, col_obs).value = "VALOR INVALIDO"
                ws_pedidos.cell(row, col_comp_total).value = "SIN TOTAL" if total_ext in (None, 0) else "DIFERENTE"
                continue

            valor_kilo = _as_int(raw_valor_kilo)
            ws_pedidos.cell(row, col_valor_kilo).value = valor_kilo

            flete_base = peso * valor_kilo
            descuento = flete_base // 4
            flete_con_desc = flete_base - descuento

            manejo = unidades * COSTO_MANEJO_POR_UNIDAD
            seguro = declarado // 200 if declarado > UMBRAL_DECLARADO_SEGURO else 0

            total_calc = flete_con_desc + manejo + seguro

            ws_pedidos.cell(row, col_prefac).value = total_calc
            ws_pedidos.cell(row, col_obs).value = "OK"

            if col_total is None or ws_pedidos.cell(row, col_total).value in (None, ""):
                ws_pedidos.cell(row, col_comp_total).value = "SIN TOTAL"
            else:
                dif = abs(total_calc - total_ext)
                if dif == 0:
                    ws_pedidos.cell(row, col_comp_total).value = "IGUAL"
                elif dif <= UMBRAL_SIMILAR:
                    ws_pedidos.cell(row, col_comp_total).value = "SIMILAR (+/- 5.000)"
                else:
                    ws_pedidos.cell(row, col_comp_total).value = "DIFERENTE"

        else:
            ws_pedidos.cell(row, col_obs).value = "NO SE ENCONTRO TARIFA"
            if col_total is None or ws_pedidos.cell(row, col_total).value in (None, ""):
                ws_pedidos.cell(row, col_comp_total).value = "SIN TOTAL"
            else:
                ws_pedidos.cell(row, col_comp_total).value = "DIFERENTE"

    wb_pedidos.save(SALIDA)
    print("Prefacturación terminada")



# TARIFAS PAQUETE


TARIFAS_PAQUETERIA = {

"URBANO":[
{"min":1,"max":3,"tarifa":5467},
{"min":4,"max":5,"tarifa":8971},
{"min":6,"max":8,"tarifa":12143}
],

"REGIONAL":[
{"min":1,"max":3,"tarifa":7794},
{"min":4,"max":5,"tarifa":11146},
{"min":6,"max":8,"tarifa":14770}
],

"NACIONAL":[
{"min":1,"max":3,"tarifa":9878},
{"min":4,"max":5,"tarifa":14500},
{"min":6,"max":8,"tarifa":17761}
],

"REEXPEDIDO":[
{"min":1,"max":3,"tarifa":27428},
{"min":4,"max":5,"tarifa":36490},
{"min":6,"max":8,"tarifa":45067}
]

}


def buscar_tarifa_paqueteria(trayecto,peso,destino):

    trayecto = normalizar(trayecto)

 

    if trayecto not in TARIFAS_PAQUETERIA:
        return None

    for escala in TARIFAS_PAQUETERIA[trayecto]:
        if escala["min"] <= peso <= escala["max"]:
            return escala["tarifa"]

    return None


# PROCESO PAQUETE

def prefacturar_paquete():

    wb = load_workbook(ARCHIVO_PEDIDOS)
    ws = wb["PAQUETE"]

    headers = {
        str(ws.cell(1,c).value).strip().upper():c
        for c in range(1,ws.max_column+1)
        if ws.cell(1,c).value
    }

    col_trayecto = headers["TRAYECTO"]
    col_destino = headers["DESTINO"]
    col_peso = headers["PESO FACTURADO"]
    col_decl = headers["DECLARADO"]
    col_total = headers["TOTAL"]

    base_col = ws.max_column

    col_tarifa = base_col+1
    col_calc = base_col+2
    col_estado = base_col+3

    ws.cell(1,col_tarifa).value="TARIFA_LISTA"
    ws.cell(1,col_calc).value="FLETE_CALCULADO"
    ws.cell(1,col_estado).value="ESTADO_CONCILIACION"

    for row in range(2,ws.max_row+1):

        trayecto = ws.cell(row,col_trayecto).value
        destino = ws.cell(row,col_destino).value

        peso = _as_int(ws.cell(row,col_peso).value)
        declarado = _as_int(ws.cell(row,col_decl).value)
        total_factura = _as_int(ws.cell(row,col_total).value)

        tarifa = buscar_tarifa_paqueteria(trayecto,peso,destino)

        if tarifa is None:
            ws.cell(row,col_estado).value="SIN TARIFA"
            continue

        ws.cell(row,col_tarifa).value = tarifa

        prima = 0
        if declarado > 10000:
            prima = int(declarado * 0.01)

        total_calc = tarifa + prima

        ws.cell(row,col_calc).value = total_calc

        if total_calc == total_factura:
            estado="IGUAL"
        else:
            estado="DIFERENTE"

        ws.cell(row,col_estado).value = estado

    wb.save(SALIDA)
    print("Prefacturación PAQUETE terminada")

# CONFIGURACIÓN DE RUTAS

NOMBRE_HOJA_DOC = "DOCUMENTO"  


# DICCIONARIO DE TARIFAS DOCUMENTO

TARIFAS_DOCUMENTO = {
    ("DE", "URBANO"): [{"valor": 3862, "adic": 0}],
    ("RF", "URBANO"): [{"valor": 7794, "adic": 1557}],
    ("DE", "NACIONAL"): [
        {"valor": 8956, "adic": 0}, {"valor": 5619, "adic": 0}, 
        {"valor": 28757, "adic": 0}, {"valor": 13171, "adic": 0}, {"valor": 23219, "adic": 0}
    ],
    ("RF", "NACIONAL"): [{"valor": 17308, "adic": 3196}],
    ("DE", "REEXPEDIDO"): [{"valor": 27428, "adic": 0},{"valor": 23219, "adic": 0}],
    ("DE", "REGIONAL"): [{"valor": 5012, "adic": 0}, {"valor": 7370, "adic": 0}],
    ("RF", "REGIONAL"): [{"valor": 10222, "adic": 0}],
    ("RF", "REEXPEDIDO"): [{"valor": 27428, "adic": 0}],
    
}

def normalizar(texto):
    if texto is None: return ""
    texto = str(texto).strip().upper()
    return unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")

def _as_int(x):
    if isinstance(x, (int, float)): return int(x)
    try:
        s = str(x or "0").strip().replace(".", "").replace(",", "").replace(" ", "")
        return int(float(s))
    except: return 0


# LÓGICA DE PROCESAMIENTO REFORZADA

def procesar_logica_documento(servicio, trayecto, peso, flete_factura):
    # 1. Limpieza de Servicio (D.E o DE -> DE)
    s_limpio = normalizar(servicio).replace(".", "").replace(" ", "")
    
    # 2. Limpieza de Trayecto y Manejo de Reexpedidos
    t_raw = normalizar(trayecto)
    
    # Lógica inteligente: Si el trayecto es REEXPEDIDO o variaciones de NOTIFICACIÓN
    if "REEXPEDIDO" in t_raw or "NOTIFIC" in t_raw:
        t_final = "REEXPEDIDO"
    elif "NACIONAL" in t_raw: t_final = "NACIONAL"
    elif "URBANO" in t_raw: t_final = "URBANO"
    elif "REGIONAL" in t_raw: t_final = "REGIONAL"
    else: t_final = t_raw

    opciones = TARIFAS_DOCUMENTO.get((s_limpio, t_final))
    if not opciones: return None

    # 3. Selección de la mejor tarifa (especialmente para Nacional)
    mejor_calc = None
    min_dif = float('inf')
    for opt in opciones:
        # Cálculo: Flete base + adicionales por peso
        calc = opt["valor"] + (max(0, peso - 1) * opt.get("adic", 0))
        dif = abs(calc - flete_factura)
        if dif < min_dif:
            min_dif = dif
            mejor_calc = calc
    return mejor_calc


# FUNCIÓN DE EJECUCIÓN PARA EXCEL
def prefacturar_documento():
    print(f"Abriendo {ARCHIVO_PEDIDOS} para conciliación contra TOTAL...")
    try:
        wb = load_workbook(ARCHIVO_PEDIDOS)
        # Soporta nombres con o sin espacio al final
        nombre_hoja = "DOCUMENTO " if "DOCUMENTO " in wb.sheetnames else "DOCUMENTO"
        
        if nombre_hoja not in wb.sheetnames:
            print(f"Error: No se encontró la hoja DOCUMENTO")
            return
            
        ws = wb[nombre_hoja]
        
        # Mapeo automático de columnas (Leemos las cabeceras del Excel)
        headers = {str(ws.cell(1, c).value).strip().upper(): c for c in range(1, ws.max_column + 1)}
        
        # Validar que exista la columna TOTAL
        if "TOTAL" not in headers:
            print("Error: No se encontró la columna 'TOTAL' en la hoja DOCUMENTO")
            return

        # Definir columnas de salida
        col_calc = headers.get("CALCULO_DOCUMENTO", ws.max_column + 1)
        col_esta = headers.get("ESTADO_CONCILIACION", ws.max_column + 2)
        
        ws.cell(1, col_calc).value = "CALCULO_DOCUMENTO"
        ws.cell(1, col_esta).value = "ESTADO_CONCILIACION"

        for row in range(2, ws.max_row + 1):
            # 1. LEER DATOS
            serv = ws.cell(row, headers["SERVICIO"]).value
            tray = ws.cell(row, headers["TRAYECTO"]).value
            peso = _as_int(ws.cell(row, headers["PESO"]).value)
            
            # LEER COLUMNA TOTAL (Aquí estaba el cambio que pediste)
            valor_total_excel = _as_int(ws.cell(row, headers["TOTAL"]).value) 

            # 2. CALCULAR SEGÚN DICCIONARIO
            resultado = procesar_logica_documento(serv, tray, peso, valor_total_excel)

            if resultado is None:
                ws.cell(row, col_esta).value = "SIN TARIFA"
                continue

            # 3. GUARDAR EL CÁLCULO
            ws.cell(row, col_calc).value = resultado
            
            # 4. COMPARACIÓN BLINDADA (Entero vs Entero)
            # Forzamos a que ambos sean números enteros para que 17308 == 17308 siempre.
            val_calculado = int(round(float(resultado)))
            val_total_real = int(round(float(valor_total_excel)))

            if val_calculado == val_total_real:
                ws.cell(row, col_esta).value = "IGUAL"
            else:
                ws.cell(row, col_esta).value = "DIFERENTE"

        wb.save(ARCHIVO_PEDIDOS)
        print("¡Procesado exitosamente comparando contra la columna TOTAL!")

    except Exception as e:
        print(f"Error en el proceso de DOCUMENTO: {e}")

